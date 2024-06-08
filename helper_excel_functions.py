# Helper excel functions
# Imports
import os 
import openpyxl
from openpyxl.styles import Font,PatternFill



# Excel formatting functions 
def launch_excel(df,name) :
    df.to_excel(f'{name}.xlsx',index=False)
    format_df(f'{name}.xlsx')
    os.startfile(f'{name}.xlsx')

def format_df(file) :
    workbook = openpyxl.load_workbook(file)
    for worksheet in workbook:
        font = Font(color='FFFFFF', bold=True)
        fill = PatternFill(start_color='5552A2', end_color='5552A2', fill_type='solid')
        for cell in worksheet[1]:
            cell.font = font
            cell.fill = fill
        for column in worksheet.columns:
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = 30      
    workbook.save(file)